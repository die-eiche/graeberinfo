#!/usr/bin/env python3
"""Liest Zulaessige Werte.xlsx und schreibt src/data/allowedValues.json."""
from __future__ import annotations

import json
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as e:
    raise SystemExit("Bitte openpyxl installieren: pip install openpyxl") from e

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "Zulaessige Werte.xlsx"
OUT = ROOT / "src" / "data" / "allowedValues.json"

SHEETS = {
    "urnen": "Gültige Urnen",
    "bestatter": "Gültige Bestatter",
    "graeber": "Gültige Gräber",
}


def unique_col(ws) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for (value,) in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        if value is None:
            continue
        text = str(value).strip()
        if not text or text in seen:
            continue
        seen.add(text)
        out.append(text)
    return out


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"Excel fehlt: {XLSX}")
    wb = load_workbook(XLSX, data_only=True)
    missing = [name for name in SHEETS.values() if name not in wb.sheetnames]
    if missing:
        raise SystemExit(f"Arbeitsblätter fehlen: {missing}; vorhanden: {wb.sheetnames}")

    data = {
        "source": XLSX.name,
        "sheets": SHEETS,
        "urnen": unique_col(wb[SHEETS["urnen"]]),
        "bestatter": unique_col(wb[SHEETS["bestatter"]]),
        "graeber": unique_col(wb[SHEETS["graeber"]]),
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(
        f"OK → {OUT} "
        f"(Urnen={len(data['urnen'])}, Bestatter={len(data['bestatter'])}, "
        f"Gräber={len(data['graeber'])})"
    )


if __name__ == "__main__":
    main()
